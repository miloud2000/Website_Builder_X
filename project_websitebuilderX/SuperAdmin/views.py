from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.template.loader import render_to_string
from django.contrib.sites.shortcuts import get_current_site
from django.core.mail import EmailMessage, send_mail
from django.utils.html import strip_tags
from django.utils.safestring import mark_safe


from websitebuilder.models import *

from websitebuilder.forms import (  
    ClienteForm,
    AdministrateurForm,
    SupportTechniqueForm,
    GestionnaireComptesForm,
    DemandeRechargerForm,
    AdditionalInfoForm,
    ClienteUpdateForm,
    ClientePasswordChangeForm,
    CommercialForm,
    UpdateAdministrateurForm,
    UpdateSupportTechniqueForm,
    UpdateGestionnaireComptesForm,
    ClienteUpdateFormSuperAdmin,
    CommercialUpdateForm,
    WebsiteForm,
    SupportForm,
)

from websitebuilder.decorators import (  
    notLoggedUsers,
    allowedUsers,
    forAdmins,
    user_not_authenticated,
    anonymous_required,
)

from websitebuilder.tokens import account_activation_token  





@anonymous_required
def home2(request):
    return render(request, "websitebuilder/home2.html")


# def home(request):  
#     if request.user.is_authenticated:
#         is_Cliente = request.user.groups.filter(name='Cliente').exists()
#         is_SupportTechnique = request.user.groups.filter(name='SupportTechnique').exists()
#         is_Administrateur = request.user.groups.filter(name='Administrateur').exists()
#     else: 
#         is_Cliente= False  
#         is_SupportTechnique= False 
#         is_Administrateur= False  
           
#     context = {"is_Cliente": is_Cliente,"is_SupportTechnique":is_SupportTechnique,"is_Administrateur":is_Administrateur}
#     return render(request, "websitebuilder/home.html",context)








def AllWebsites_client_status(request):
    per_page = int(request.GET.get('per_page', 10))
    page_number = request.GET.get('page')

    all_status = Websites_client_statu.objects.all().order_by('-date_created')

    paginator = Paginator(all_status, per_page)
    page_obj = paginator.get_page(page_number)

    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        "page_obj": page_obj,
        "AllWebsites_client_status": page_obj.object_list,
        "per_page": per_page,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, "SuperAdmin/AllMergedWebsiteBuilder.html", context)




#SuperAdmin


# #Home SuperAdmin
@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def homeSuperAdmin(request):  
    return render(request, "SuperAdmin/homeSuperAdmin.html")


from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth
from websitebuilder.models import AchatWebsites
from collections import OrderedDict
import calendar


#DashbordHome SuperAdmin
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import timedelta
from collections import OrderedDict
import calendar

from websitebuilder.models import Websites, Supports, Cliente, AchatWebsites
from django.contrib.auth.models import User
import json
from django.core.paginator import Paginator


from django.utils.timesince import timesince
from django.utils.translation import activate
from django.urls import reverse
import re

def get_superadmin_notifications():
    activate('fr')

    actions = HistoriqueAction.objects.select_related('utilisateur').order_by('-date')[:10]

    notifications = []
    for action in actions:
        notif = {
            'message': f"{action.action} sur {action.objet}",
            'details': f"Par {action.utilisateur.username}" if action.utilisateur else "Par Utilisateur inconnu",
            'time': timesince(action.date) + " il y a",
            'icon': 'fe-activity',
            'color': 'warning',
        }

        match = re.search(r'Demande ID\s+#(\d+)', action.details or '')
        if match:
            demande_id = match.group(1)
            try:
                demande = DemandeRecharger.objects.get(id=demande_id)
                notif['url'] = reverse('detail_demande_recharge_superadmin', args=[demande.id])
            except DemandeRecharger.DoesNotExist:
                pass
            
        
        match_support = re.search(r'«\s*(.*?)\s*»', action.details or '')
        if match_support and 'DemandeSupport' in action.objet:
            code_support = match_support.group(1)
            try:
                demande = DemandeSupport.objects.get(code_DemandeSupport=code_support)
                notif['url'] = reverse('detail_demande_support_superadmin', args=[demande.id])
            except DemandeSupport.DoesNotExist:
                pass

        notifications.append(notif)

    return notifications






def get_all_ticket_messages():
    tickets = Ticket.objects.filter(
        conversations__isnull=False
    ).distinct().order_by('-date_updated')[:10]

    messages = []

    for ticket in tickets:
        last_convo = ticket.conversations.order_by('-timestamp').first()
        if last_convo:
            if last_convo.sender_type == 'Cliente':
                image_path = 'assets/images/R.png'
            elif last_convo.sender_type == 'SupportTechnique':
                image_path = 'assets/images/R.png'
            elif last_convo.sender_type == 'GestionnaireComptes':
                image_path = 'assets/images/R.png'
            else:
                image_path = 'assets/images/faces/R.png'

            messages.append({
            'sender': last_convo.sender,
            'title': f"Ticket : {ticket.typeTicket}",
            'subtitle': f"Statut : {ticket.status}",
            'time': timesince(ticket.date_updated),
            'image': image_path,
            'ticket_id': ticket.id, 
        })


    return messages





@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def dashbordHomeSuperAdmin(request):
    today = timezone.now()
    start_week = today - timedelta(days=today.weekday())
    start_last_week = start_week - timedelta(days=7)
    end_last_week = start_week - timedelta(seconds=1)

    clients_this_week = User.objects.filter(groups__name='Cliente', date_joined__gte=start_week).count()
    clients_last_week = User.objects.filter(groups__name='Cliente', date_joined__range=(start_last_week, end_last_week)).count()
    client_growth = ((clients_this_week - clients_last_week) / clients_last_week * 100) if clients_last_week > 0 else (-100.0 if clients_this_week == 0 else 100.0)

    start_6_days_ago = today - timedelta(days=6)
    start_12_days_ago = today - timedelta(days=12)
    websites_now = Websites.objects.filter(date_created__gte=start_6_days_ago).count()
    websites_before = Websites.objects.filter(date_created__range=(start_12_days_ago, start_6_days_ago)).count()
    websites_growth = ((websites_now - websites_before) / websites_before * 100) if websites_before > 0 else (-100.0 if websites_now == 0 else 100.0)

    start_9_days_ago = today - timedelta(days=9)
    start_18_days_ago = today - timedelta(days=18)
    supports_now = Supports.objects.filter(date_created__gte=start_9_days_ago).count()
    supports_before = Supports.objects.filter(date_created__range=(start_18_days_ago, start_9_days_ago)).count()
    supports_growth = ((supports_now - supports_before) / supports_before * 100) if supports_before > 0 else (-100.0 if supports_now == 0 else 100.0)

    start_year = today.replace(month=1, day=1)
    solde_this_year = Cliente.objects.filter(date_created__gte=start_year).aggregate(Sum('solde'))['solde__sum'] or 0
    solde_before = Cliente.objects.filter(date_created__lt=start_year).aggregate(Sum('solde'))['solde__sum'] or 0
    solde_growth = ((solde_this_year - solde_before) / solde_before * 100) if solde_before > 0 else (-100.0 if solde_this_year == 0 else 100.0)

    total_sales = AchatWebsites.objects.aggregate(Sum('prix_achat'))['prix_achat__sum'] or 0
    total_orders = AchatWebsites.objects.count()
    delivered_orders = AchatWebsites.objects.filter(BuilderStatus='Builder').count()
    cancelled_orders = AchatWebsites.objects.filter(BuilderStatus='Not yet').count()

   
    
    latest_recharges = DemandeRecharger.objects.select_related('cliente__user').order_by('-date_created')[:6]
    
    total_achats = AchatWebsites.objects.count()
    
    website_achat_counts = (
    AchatWebsites.objects
    .values('websites__name')
    .annotate(count=Count('id'))
    )
    
    website_achat_percentages = []
    for entry in website_achat_counts:
        name = entry['websites__name']
        count = entry['count']
        percentage = round((count / total_achats) * 100, 2) if total_achats > 0 else 0
        website_achat_percentages.append({
            'name': name,
            'percentage': percentage
        })
        
        
    total_achat_supports = AchatSupport.objects.count()
    
    support_achat_counts = (
    AchatSupport.objects
    .values('support__name')
    .annotate(count=Count('id'))
    )
    
    support_achat_percentages = []
    for entry in support_achat_counts:
        name = entry['support__name']
        count = entry['count']
        percentage = round((count / total_achat_supports) * 100, 2) if total_achat_supports > 0 else 0
        support_achat_percentages.append({
            'name': name,
            'count': count,
            'percentage': percentage
        })
    
    
    clients = Cliente.objects.select_related('user').all()
    
    search_query = request.GET.get('search', '')
    per_page = request.GET.get('per_page', 10)

    clients = Cliente.objects.select_related('user')

    if search_query:
        clients = clients.filter(
            Q(nom__icontains=search_query) |
            Q(prenom__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    paginator = Paginator(clients, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
        
        
    context = {
    'clients_this_week': clients_this_week,
    'client_growth': round(client_growth, 2),
    'websites_growth': round(websites_growth, 2),
    'supports_growth': round(supports_growth, 2),
    'solde_growth': round(solde_growth, 2),
    'total_website': Websites.objects.count(),
    'total_service': Supports.objects.count(),
    'total_solde': round(Cliente.objects.aggregate(Sum('solde'))['solde__sum'] or 0, 2),

    # Graph data for each chart

    'latest_recharges': latest_recharges,
    'website_achat_percentages': website_achat_percentages,
    'support_achat_percentages': support_achat_percentages,
    
    'clients': clients,
    
    'page_obj': page_obj,
    'search_query': search_query,
    'per_page': int(per_page),
    
    'notifications': notifications,
    'messages_dropdown': messages_dropdown,
}

    return render(request, "SuperAdmin/dashbordHomeSuperAdmin.html", context)






@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def detail_SuperAdmin(request):  
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()

    historique_actions = HistoriqueAction.objects.select_related('utilisateur').order_by('-date')[:10]

    context = {
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
        'historique_actions': historique_actions,
    }
    return render(request, "SuperAdmin/detail_SuperAdmin.html", context)








@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def addAdministrateur(request):
    if request.method == 'POST':
        form = AdministrateurForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.save()

            Administrateur.objects.create(
                user=user,
                name=form.cleaned_data['name'],
                email=form.cleaned_data['email'],
                phone=form.cleaned_data['phone'],
            )

            group = Group.objects.get(name="Administrateur")
            user.groups.add(group)

            messages.success(request, f"{user.username} a été créé avec succès !")
            return redirect('AdministrateurSuperAdmin')
        else:
            messages.error(request, "Formulaire invalide. Veuillez corriger les erreurs.")
    else:
        form = AdministrateurForm()
        
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()

    context = {'form': form,
               'notifications': notifications,
            'messages_dropdown': messages_dropdown,
        }
    return render(request, 'SuperAdmin/addAdministrateur.html', context)




from django.db.models import Q

#Superadmin can show all Administrateur
@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def AdministrateurSuperAdmin(request): 
    query = request.GET.get('q')
    status_filter = request.GET.get('status')
    per_page = request.GET.get('per_page', 10)

    Administrateurs = Administrateur.objects.all()

    if query:
        Administrateurs = Administrateurs.filter(
            Q(user__username__icontains=query) |
            Q(name__icontains=query) |
            Q(email__icontains=query) |
            Q(phone__icontains=query)
        )

    if status_filter:
        Administrateurs = Administrateurs.filter(Status__icontains=status_filter)

    paginator = Paginator(Administrateurs, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'Administrateurs': Administrateurs,
         'page_obj': page_obj,
        'query': query,
        'status_filter': status_filter,
        'per_page': int(per_page),
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
        }
    return render(request, "SuperAdmin/AdministrateurSuperAdmin.html", context)




from django.core.paginator import Paginator

@login_required
@allowedUsers(allowedGroups=['SuperAdmin'])
def historique_administrateur(request, admin_id):
    administrateur = get_object_or_404(Administrateur, id=admin_id)
    actions = HistoriqueAction.objects.filter(utilisateur=administrateur.user).order_by('-date')

    # Filtres
    query = request.GET.get('q')
    action_filter = request.GET.get('action')
    objet_filter = request.GET.get('objet')
    date_min = request.GET.get('date_min')
    date_max = request.GET.get('date_max')
    per_page = int(request.GET.get('per_page', 10))  # valeur par défaut = 10

    if query:
        actions = actions.filter(
            Q(action__icontains=query) |
            Q(objet__icontains=query) |
            Q(details__icontains=query)
        )
    if action_filter:
        actions = actions.filter(action=action_filter)
    if objet_filter:
        actions = actions.filter(objet=objet_filter)
    if date_min:
        actions = actions.filter(date__date__gte=date_min)
    if date_max:
        actions = actions.filter(date__date__lte=date_max)

    # Pagination
    paginator = Paginator(actions, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Valeurs uniques pour les dropdowns
    action_choices = HistoriqueAction.objects.values_list('action', flat=True).distinct()
    objet_choices = HistoriqueAction.objects.values_list('objet', flat=True).distinct()

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'administrateur': administrateur,
        'page_obj': page_obj,
        'actions': page_obj.object_list,
        'action_choices': action_choices,
        'objet_choices': objet_choices,
        'per_page': per_page,
        'query': query,
        'action_filter': action_filter,
        'objet_filter': objet_filter,
        'date_min': date_min,
        'date_max': date_max,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/historique_administrateur.html', context)





from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    BaseDocTemplate, Frame, PageTemplate,
    Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet

class PDFWithHeaderFooter(BaseDocTemplate):
    def __init__(self, filename, admin_name, **kwargs):
        super().__init__(filename, **kwargs)
        self.admin_name = admin_name

        # Cadre principal (tout sauf footer)
        frame = Frame(
            self.leftMargin,
            self.bottomMargin,
            self.width,
            self.height - 2*cm,  # espace pour footer
            id='normal'
        )
        self.addPageTemplates([
            PageTemplate(
                id='Standard',
                frames=frame,
                onPage=self._header,
                onPageEnd=self._footer
            )
        ])

    def _header(self, canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica-Bold', 14)
        canvas.drawString(2*cm, A4[1] - 2*cm,
                          f"Historique des actions de {self.admin_name}")
        canvas.restoreState()

    def _footer(self, canvas, doc):
        canvas.saveState()
        page_num = f"Page {canvas.getPageNumber()}"
        canvas.setFont('Helvetica', 9)
        canvas.drawRightString(
            A4[0] - 2*cm, 1.5*cm, page_num
        )
        canvas.restoreState()



@login_required
@allowedUsers(allowedGroups=['SuperAdmin'])
def export_historique_pdf(request, admin_id):
    administrateur = get_object_or_404(Administrateur, id=admin_id)
    actions = HistoriqueAction.objects.filter(
        utilisateur=administrateur.user
    ).order_by('-date')

    # Préparation de la réponse PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="historique_{administrateur.user.username}.pdf"'
    )

    # Création du document
    doc = PDFWithHeaderFooter(
        filename=response,
        admin_name=administrateur.name,
        pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=3*cm, bottomMargin=2*cm
    )
    styles = getSampleStyleSheet()
    story = []

    # Petite marge sous l'en-tête
    story.append(Spacer(1, 1*cm))

    # Construction des données du tableau
    data = [['Date', 'Action', 'Objet', 'Détails']]
    for a in actions:
        data.append([
            a.date.strftime('%d/%m/%Y %H:%M'),
            a.action,
            a.objet,
            a.details or ''
        ])

    # Création et style du tableau
    table = Table(data, colWidths=[3*cm, 4*cm, 4*cm, 6*cm], repeatRows=1)
    table.setStyle(TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e9ecef')),
        ('FONTNAME',    (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0), 11),
        ('ALIGN',       (0, 0), (-1, 0), 'CENTER'),

        # Bordures et alignement
        ('GRID',        (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN',      (0, 0), (-1, -1), 'TOP'),

        # Zebra-striping automatique
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7f7f7')]),
    ]))

    story.append(table)
    doc.build(story)
    return response






import csv
from django.http import HttpResponse

@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def export_admins_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="administrateurs.csv"'

    writer = csv.writer(response)
    writer.writerow(['Username', 'Email', 'Phone', 'Status', 'Date ajout'])

    for admin in Administrateur.objects.all():
        writer.writerow([
            admin.user.username,
            admin.email,
            admin.phone,
            admin.Status,
            admin.date_created.strftime('%d/%m/%Y %H:%M')
        ])

    return response


import openpyxl
from django.http import HttpResponse

@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def export_admins_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Administrateurs"

    headers = ['Username', 'Email', 'Phone', 'Status', 'Date ajout']
    ws.append(headers)

    for admin in Administrateur.objects.all():
        ws.append([
            admin.user.username,
            admin.email,
            admin.phone,
            admin.Status,
            admin.date_created.strftime('%d/%m/%Y %H:%M')
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="administrateurs.xlsx"'
    wb.save(response)
    return response


from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO

@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def export_admins_pdf(request):
    admins = Administrateur.objects.all()
    template = get_template('SuperAdmin/pdf_admins_template.html')
    html = template.render({'admins': admins})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="administrateurs.pdf"'

    pisa_status = pisa.CreatePDF(BytesIO(html.encode('utf-8')), dest=response)
    return response




@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def updateAdministrateur(request, pk):
    administrateur = get_object_or_404(Administrateur, id=pk)
    form = UpdateAdministrateurForm(instance=administrateur)

    if request.method == 'POST':
        form = UpdateAdministrateurForm(request.POST, instance=administrateur)
        if form.is_valid():
            form.save()
            messages.success(request, "Administrateur mis à jour avec succès.")
            return redirect('AdministrateurSuperAdmin')
        else:
            messages.error(request, "Formulaire invalide. Veuillez corriger les erreurs.")
            
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
            'form': form,
            'notifications': notifications,
            'messages_dropdown': messages_dropdown,
        }
    return render(request, 'SuperAdmin/updateAdministrateur.html', context)




@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def deleteAdministrateur(request, pk):
    administrateur = get_object_or_404(Administrateur, id=pk)
    user = administrateur.user

    if request.method == 'POST':
        user.delete() 
        messages.success(request, "Administrateur supprimé avec succès.")
        return redirect('AdministrateurSuperAdmin')

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'administrateur': administrateur,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/deleteAdministrateur.html', context)



#The SuperAdmin can add a SupportTechnique
@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def addSupportTechnique(request):
    form = SupportTechniqueForm()
    if request.method == 'POST':
        form = SupportTechniqueForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)  
            user.save()

            SupportTechnique.objects.create(
                user=user,
                name=form.cleaned_data.get('name'),
                email=form.cleaned_data.get('email'),
                phone=form.cleaned_data.get('phone'),
            )

            # Add user to the 'SupportTechnique' group
            group = Group.objects.get(name="SupportTechnique")
            user.groups.add(group)

            messages.success(request, f"{user.username} created successfully!")
            
            return redirect('SupportTechniqueSuperAdmin')
        else:
            messages.error(request, "Invalid form submission. Please correct the errors below.")  
    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
        
    context = {
        'form': form,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
        }
    return render(request, 'SuperAdmin/addSupportTechnique.html', context)





@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def updateSupportTechnique(request, pk):
    support = get_object_or_404(SupportTechnique, id=pk)
    form = UpdateSupportTechniqueForm(instance=support)

    if request.method == 'POST':
        form = UpdateSupportTechniqueForm(request.POST, instance=support)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Support Technique mis à jour avec succès.")
            return redirect('SupportTechniqueSuperAdmin')
        
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()

    context = {
        'form': form,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
        }
    return render(request, 'SuperAdmin/updateSupportTechnique.html', context)



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def deleteSupportTechnique(request, pk):
    support = get_object_or_404(SupportTechnique, id=pk)
    user = support.user

    if request.method == 'POST':
        user.delete()
        messages.success(request, "🗑️ Support Technique supprimé avec succès.")
        return redirect('SupportTechniqueSuperAdmin')
    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()

    context = {
        'support': support,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
        }
    return render(request, 'SuperAdmin/deleteSupportTechnique.html', context)




from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.utils.encoding import smart_str

def export_support_history_pdf(request, support_id):
    support = SupportTechnique.objects.get(id=support_id)
    actions = HistoriqueAction.objects.filter(utilisateur__supporttechnique=support)

    template = get_template("SuperAdmin/pdf_support_history.html")
    html = template.render({"support": support, "actions": actions})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="historique_{smart_str(support.user.username)}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    return response if not pisa_status.err else HttpResponse("Erreur lors de la génération du PDF", status=500)





#Superadmin can show all SupportTechnique
from django.core.paginator import Paginator

@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def SupportTechniqueSuperAdmin(request): 
    query = request.GET.get('q')
    status_filter = request.GET.get('status')
    history_id = request.GET.get('history')
    per_page = int(request.GET.get('per_page', 10))  # valeur par défaut = 10

    supportTechniques = SupportTechnique.objects.all()

    if query and query != 'None':
        supportTechniques = supportTechniques.filter(
            Q(user__username__icontains=query) |
            Q(name__icontains=query) |
            Q(email__icontains=query) |
            Q(phone__icontains=query)
        )

    if status_filter and status_filter != 'None':
        supportTechniques = supportTechniques.filter(Status__iexact=status_filter)
    
    # Pagination
    paginator = Paginator(supportTechniques, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Historique
    history_actions = []
    if history_id:
        history_actions = HistoriqueAction.objects.filter(utilisateur__supporttechnique__id=history_id).order_by('-date')

    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'page_obj': page_obj,
        'supportTechniques': page_obj.object_list,
        'history_actions': history_actions,
        'selected_id': history_id,
        'per_page': per_page,
        'query': query,
        'status_filter': status_filter,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, "SuperAdmin/SupportTechniqueSuperAdmin.html", context)





@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def support_technique_history(request, pk):
    support = get_object_or_404(SupportTechnique, pk=pk)
    actions = HistoriqueAction.objects.filter(utilisateur=support.user).order_by('-date')

    # Filtres
    query = request.GET.get('q')
    action_filter = request.GET.get('action')
    objet_filter = request.GET.get('objet')
    date_min = request.GET.get('date_min')
    date_max = request.GET.get('date_max')
    per_page = int(request.GET.get('per_page', 10))  # valeur par défaut = 10

    if query:
        actions = actions.filter(
            Q(action__icontains=query) |
            Q(objet__icontains=query) |
            Q(details__icontains=query)
        )
    if action_filter:
        actions = actions.filter(action=action_filter)
    if objet_filter:
        actions = actions.filter(objet=objet_filter)
    if date_min:
        actions = actions.filter(date__date__gte=date_min)
    if date_max:
        actions = actions.filter(date__date__lte=date_max)

    # Pagination
    paginator = Paginator(actions, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Valeurs dynamiques pour les dropdowns
    action_choices = HistoriqueAction.objects.values_list('action', flat=True).distinct()
    objet_choices = HistoriqueAction.objects.values_list('objet', flat=True).distinct()

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'support': support,
        'page_obj': page_obj,
        'actions': page_obj.object_list,
        'action_choices': action_choices,
        'objet_choices': objet_choices,
        'per_page': per_page,
        'query': query,
        'action_filter': action_filter,
        'objet_filter': objet_filter,
        'date_min': date_min,
        'date_max': date_max,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/support_technique_history.html', context)




import csv
from django.http import HttpResponse

def export_support_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="support_technique.csv"'

    writer = csv.writer(response)
    writer.writerow(['Username', 'Email', 'Phone', 'Status', 'Date ajout'])

    for s in SupportTechnique.objects.all():
        writer.writerow([
            s.user.username,
            s.email,
            s.phone,
            s.Status,
            s.date_created.strftime('%d/%m/%Y %H:%M')
        ])

    return response




import openpyxl

def export_support_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Support Technique"

    ws.append(['Username', 'Email', 'Phone', 'Status', 'Date ajout'])

    for s in SupportTechnique.objects.all():
        ws.append([
            s.user.username,
            s.email,
            s.phone,
            s.Status,
            s.date_created.strftime('%d/%m/%Y %H:%M')
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="support_technique.xlsx"'
    wb.save(response)
    return response


from xhtml2pdf import pisa
from django.template.loader import get_template
from io import BytesIO

def export_support_pdf(request):
    supports = SupportTechnique.objects.all()
    template = get_template('SuperAdmin/pdf_support_template.html')
    html = template.render({'supports': supports})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="support_technique.pdf"'
    pisa.CreatePDF(BytesIO(html.encode('utf-8')), dest=response)
    return response




#The SuperAdmin can add a GestionnaireComptes
@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def addGestionnaireComptes(request):
    form = GestionnaireComptesForm()
    if request.method == 'POST':
        form = GestionnaireComptesForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False) 
            user.save()  

            GestionnaireComptes.objects.create(
                user=user,
                name=form.cleaned_data.get('name'),
                email=form.cleaned_data.get('email'),
                phone=form.cleaned_data.get('phone'),
            )

            # Add user to the 'GestionnaireComptes' group
            group = Group.objects.get(name="GestionnaireComptes")
            user.groups.add(group)

            messages.success(request, f"{user.username} created successfully!")
            
            return redirect('GestionnaireComptesSuperAdmin')
        else:
            messages.error(request, "Invalid form submission. Please correct the errors below.")  
    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'form': form,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/addGestionnaireComptes.html', context)



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def updateGestionnaireComptes(request, pk):
    gestionnaire = get_object_or_404(GestionnaireComptes, id=pk)
    form = UpdateGestionnaireComptesForm(instance=gestionnaire)

    if request.method == 'POST':
        form = UpdateGestionnaireComptesForm(request.POST, instance=gestionnaire)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Gestionnaire mis à jour avec succès.")
            return redirect('GestionnaireComptesSuperAdmin')

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'form': form,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/updateGestionnaireComptes.html', context)




@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def deleteGestionnaireComptes(request, pk):
    gestionnaire = get_object_or_404(GestionnaireComptes, pk=pk)

    if request.method == 'POST':
        if gestionnaire.user:
            gestionnaire.user.delete()
        gestionnaire.delete()
        messages.success(request, "✅ Le gestionnaire a été supprimé avec succès.")
        return redirect('GestionnaireComptesSuperAdmin')

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'gestionnaire': gestionnaire,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/deleteGestionnaireComptes.html', context)




@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def GestionnaireComptesSuperAdmin(request):
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')
    per_page = int(request.GET.get('per_page', 10))
    page_number = request.GET.get('page')

    gestionnaires = GestionnaireComptes.objects.all()

    if query:
        gestionnaires = gestionnaires.filter(
            models.Q(name__icontains=query) |
            models.Q(email__icontains=query) |
            models.Q(phone__icontains=query) |
            models.Q(user__username__icontains=query)
        )
    if status:
        gestionnaires = gestionnaires.filter(Status=status)

    paginator = Paginator(gestionnaires.order_by('-id'), per_page)
    page_obj = paginator.get_page(page_number)

    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'GestionnairesComptes': page_obj.object_list,
        'page_obj': page_obj,
        'query': query,
        'status_filter': status,
        'per_page': per_page,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, "SuperAdmin/GestionnaireComptesSuperAdmin.html", context)



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def gestionnaire_comptes_history(request, pk):
    gestionnaire = get_object_or_404(GestionnaireComptes, pk=pk)
    actions = HistoriqueAction.objects.filter(utilisateur=gestionnaire.user).order_by('-date')

    # Filtres
    action_filter = request.GET.get('action')
    objet_filter = request.GET.get('objet')
    date_min = request.GET.get('date_min')
    date_max = request.GET.get('date_max')
    per_page = int(request.GET.get('per_page', 10))
    page_number = request.GET.get('page')

    if action_filter:
        actions = actions.filter(action=action_filter)

    if objet_filter:
        actions = actions.filter(objet=objet_filter)

    if date_min:
        actions = actions.filter(date__date__gte=date_min)

    if date_max:
        actions = actions.filter(date__date__lte=date_max)

    # Pagination
    paginator = Paginator(actions, per_page)
    page_obj = paginator.get_page(page_number)

    # Dropdowns dynamiques
    action_choices = HistoriqueAction.objects.values_list('action', flat=True).distinct()
    objet_choices = HistoriqueAction.objects.values_list('objet', flat=True).distinct()

    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'gestionnaire': gestionnaire,
        'actions': page_obj.object_list,
        'page_obj': page_obj,
        'per_page': per_page,
        'action_choices': action_choices,
        'objet_choices': objet_choices,
        'action_filter': action_filter,
        'objet_filter': objet_filter,
        'date_min': date_min,
        'date_max': date_max,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, "SuperAdmin/gestionnaire_comptes_history.html", context)



from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
import io

@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def gestionnaire_history_pdf(request, id):
    gestionnaire = get_object_or_404(GestionnaireComptes, id=id)
    actions = HistoriqueAction.objects.filter(utilisateur=gestionnaire.user).order_by('-date')

    template_path = 'SuperAdmin/gestionnaire_history_pdf.html'
    context = {
        'gestionnaire': gestionnaire,
        'actions': actions,
    }

    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename="historique_{gestionnaire.user.username}.pdf"'

    pisa_status = pisa.CreatePDF(io.StringIO(html), dest=response)

    if pisa_status.err:
        return HttpResponse('Erreur lors de la génération du PDF', status=500)
    return response




@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def export_gestionnaire_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="gestionnaires.csv"'
    writer = csv.writer(response)
    writer.writerow(['Username', 'Name', 'Email', 'Phone', 'Status', 'Date Created'])
    for g in GestionnaireComptes.objects.all():
        writer.writerow([g.user.username, g.name, g.email, g.phone, g.Status, g.date_created])
    return response


import pandas as pd
@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def export_gestionnaire_excel(request):
    data = []
    for g in GestionnaireComptes.objects.all():
        data.append({
            'Username': g.user.username,
            'Name': g.name,
            'Email': g.email,
            'Phone': g.phone,
            'Status': g.Status,
            'Date Created': g.date_created.replace(tzinfo=None), 
        })

    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="gestionnaires.xlsx"'
    df.to_excel(response, index=False)
    return response



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def export_gestionnaire_pdf(request):
    gestionnaires = GestionnaireComptes.objects.all()
    html = render_to_string("SuperAdmin/pdf_gestionnaire_template.html", {'gestionnaires': gestionnaires})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="gestionnaires.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Erreur lors de la génération du PDF")
    return response






@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def ClienteSuperAdmin(request):
    query = request.GET.get('q', '')
    per_page = int(request.GET.get('per_page', 10))
    page_number = request.GET.get('page')

    clientes = Cliente.objects.all()

    if query:
        clientes = clientes.filter(
            models.Q(nom__icontains=query) |
            models.Q(prenom__icontains=query) |
            models.Q(email__icontains=query) |
            models.Q(phone__icontains=query) |
            models.Q(user__username__icontains=query) |
            models.Q(code_client__icontains=query)
        )

    paginator = Paginator(clientes.order_by('-id'), per_page)
    page_obj = paginator.get_page(page_number)

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'clientes': page_obj.object_list,
        'page_obj': page_obj,
        'query': query,
        'per_page': per_page,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, "SuperAdmin/ClienteSuperAdmin.html", context)




#The SuperAdmin can add a clientes
@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def addCliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            # Create a new user
            user = User.objects.create_user(
                username=form.cleaned_data.get('username'),
                email=form.cleaned_data.get('email'),
                password=form.cleaned_data.get('password1')
            )

            # Create a new Cliente
            Cliente.objects.create(
                user=user,
                prenom=form.cleaned_data.get('prenom'),
                nom=form.cleaned_data.get('nom'),
                email=form.cleaned_data.get('email'),
                phone=form.cleaned_data.get('phone'),
            )

            # Add user to the 'Cliente' group
            group = Group.objects.get(name="Cliente")
            user.groups.add(group)

            messages.success(request, f"{user.username} created successfully!")
            return redirect('ClienteSuperAdmin')
        else:
            messages.error(request, "Invalid form submission. Please correct the errors below.")
    else:
        form = ClienteForm()
        
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'form': form,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/addCliente.html', context)




@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def updateCliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)

    if request.method == 'POST':
        form = ClienteUpdateFormSuperAdmin(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Client modifié avec succès.")
            return redirect('ClienteSuperAdmin')
        else:
            messages.error(request, "❌ Erreur dans le formulaire.")
    else:
        form = ClienteUpdateFormSuperAdmin(instance=cliente)

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'form': form,
        'cliente': cliente,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/updateCliente.html', context)



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def deleteCliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)

    if request.method == 'POST':
        if cliente.user:
            cliente.user.delete()
        cliente.delete()
        messages.success(request, "✅ Client supprimé avec succès.")
        return redirect('ClienteSuperAdmin')
    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'cliente': cliente,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/deleteCliente.html', context)



def export_clientes_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="clientes.csv"'
    writer = csv.writer(response)
    writer.writerow(['Username', 'Nom', 'Prenom', 'Email', 'Téléphone', 'Solde', 'Date'])
    for c in Cliente.objects.all():
        writer.writerow([
            c.user.username,
            c.nom,
            c.prenom,
            c.email,
            c.phone,
            c.solde,
            c.date_created.replace(tzinfo=None) if c.date_created else ''
        ])
    return response



def export_clientes_excel(request):
    data = []
    for c in Cliente.objects.all():
        date_created = c.date_created.replace(tzinfo=None) if c.date_created else ''
        data.append({
            'Username': c.user.username,
            'Nom': c.nom,
            'Prenom': c.prenom,
            'Email': c.email,
            'Téléphone': c.phone,
            'Solde': c.solde,
            'Date': date_created,
        })
    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="clientes.xlsx"'
    df.to_excel(response, index=False)
    return response



def export_clientes_pdf(request):
    clientes = Cliente.objects.all()
    html = render_to_string("SuperAdmin/pdf_cliente_template.html", {'clientes': clientes, 'now': timezone.now()})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="clientes.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Erreur lors de la génération du PDF")
    return response



from websitebuilder.models import (
    Cliente, DemandeSupport, AchatSupport,
    LocationWebsites, GetFreeWebsites
)
from .utils.dashboard_filters import filter_demandes, filter_achats, filter_tickets, filter_achat_supports
from .utils.exports import (
    export_achats_excel, export_achats_pdf,
    export_tickets_excel, export_tickets_pdf,export_achat_supports_excel, export_achat_supports_pdf
)

@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def cliente_activity_dashboard(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)

    demandes, demande_export = filter_demandes(request, cliente)
    achats, achat_export = filter_achats(request, cliente)
    tickets, ticket_export = filter_tickets(request, cliente)
    achat_supports, support_export = filter_achat_supports(request, cliente)

    if demande_export == 'excel':
        return export_demandes_excel(demandes, cliente)
    if demande_export == 'pdf':
        return export_demandes_pdf(demandes, cliente)

    if achat_export == 'excel':
        return export_achats_excel(achats, cliente)
    if achat_export == 'pdf':
        return export_achats_pdf(achats, cliente)

    if ticket_export == 'excel':
        return export_tickets_excel(tickets, cliente)
    if ticket_export == 'pdf':
        return export_tickets_pdf(tickets, cliente)

    if support_export == 'excel':
        return export_achat_supports_excel(achat_supports, cliente)
    if support_export == 'pdf':
        return export_achat_supports_pdf(achat_supports, cliente)

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'cliente': cliente,
        'demandes': demandes,
        'achats': achats,
        'tickets': tickets,
        'achat_supports': achat_supports,
        'locations': LocationWebsites.objects.filter(cliente=cliente),
        'free_websites': GetFreeWebsites.objects.filter(cliente=cliente),
        'supports': DemandeSupport.objects.filter(cliente=cliente),

        # Filtres actifs
        'demande_status': request.GET.get('demande_status'),
        'demande_date_min': request.GET.get('demande_date_min'),
        'demande_date_max': request.GET.get('demande_date_max'),
        'achat_status': request.GET.get('achat_status'),
        'achat_date_min': request.GET.get('achat_date_min'),
        'achat_date_max': request.GET.get('achat_date_max'),
        'ticket_status': request.GET.get('ticket_status'),
        'ticket_date_min': request.GET.get('ticket_date_min'),
        'ticket_date_max': request.GET.get('ticket_date_max'),
        'support_status': request.GET.get('support_status'),
        'support_conso': request.GET.get('support_conso'),
        
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/cliente_activity_dashboard.html', context)






@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def CommercialSuperAdmin(request): 
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    per_page = int(request.GET.get('per_page', 10))
    page_number = request.GET.get('page')

    commercials = Commercial.objects.all()

    if query:
        commercials = commercials.filter(
            Q(name__icontains=query) |
            Q(email__icontains=query) |
            Q(phone__icontains=query) |
            Q(user__username__icontains=query)
        )

    if status_filter:
        commercials = commercials.filter(status=status_filter)

    paginator = Paginator(commercials.order_by('-id'), per_page)
    page_obj = paginator.get_page(page_number)


    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'commercials': page_obj.object_list,
        'page_obj': page_obj,
        'query': query,
        'status_filter': status_filter,
        'per_page': per_page,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, "SuperAdmin/CommercialSuperAdmin.html", context)




import csv
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from io import BytesIO


@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def historique_commercial(request, commercial_id):
    commercial = get_object_or_404(Commercial, id=commercial_id)

    # Filtres
    action_filter = request.GET.get('action')
    objet_filter = request.GET.get('objet')
    per_page = int(request.GET.get('per_page', 10))
    page_number = request.GET.get('page')

    actions = HistoriqueAction.objects.filter(utilisateur=commercial.user).order_by('-date')

    if action_filter:
        actions = actions.filter(action=action_filter)
    if objet_filter:
        actions = actions.filter(objet=objet_filter)

    # Export Excel
    if request.GET.get('export') == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Historique"
        ws.append(['Action', 'Objet', 'Détails', 'Date'])
        for a in actions:
            ws.append([a.action, a.objet, a.details, a.date.strftime('%d/%m/%Y %H:%M')])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=historique_{commercial.user.username}.xlsx'
        wb.save(response)
        return response

    # Export PDF
    if request.GET.get('export') == 'pdf':
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"<b>Historique du commercial : {commercial.user.username}</b>", styles['Title']))
        elements.append(Spacer(1, 12))
        data = [['Date', 'Action', 'Objet', 'Détails']]
        for a in actions:
            data.append([
                a.date.strftime('%d/%m/%Y %H:%M'),
                a.action,
                a.objet,
                a.details
            ])
        table = Table(data, colWidths=[100, 150, 100, 180])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=historique_{commercial.user.username}.pdf'
        return response

    # Pagination
    paginator = Paginator(actions, per_page)
    page_obj = paginator.get_page(page_number)

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    # Choix fixes
    action_choices = [
        "Ajout d'un nouveau client",
        "Modification d'un client",
        "Suppression d'un client"
    ]
    objet_choices = [
        "Cliente",
        "Support",
        "SiteWeb"
    ]

    context = {
        'commercial': commercial,
        'actions': page_obj.object_list,
        'page_obj': page_obj,
        'per_page': per_page,
        'selected_action': action_filter,
        'selected_objet': objet_filter,
        'action_choices': action_choices,
        'objet_choices': objet_choices,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/historique_commercial.html', context)




@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def statistiques_commercial(request, commercial_id):
    commercial = get_object_or_404(Commercial, id=commercial_id)

    actions = HistoriqueAction.objects.filter(
        utilisateur=commercial.user
    )

    # ✅ Résumé par type d’action
    resume = {
        "Ajout": actions.filter(action__icontains="Ajout").count(),
        "Modification": actions.filter(action__icontains="Modification").count(),
        "Suppression": actions.filter(action__icontains="Suppression").count(),
    }

    context = {
        'commercial': commercial,
        'resume': resume,
        'total': actions.count(),
    }
    return render(request, 'SuperAdmin/statistiques_commercial.html', context)





@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def addCommercial(request):
    if request.method == 'POST':
        form = CommercialForm(request.POST)
        if form.is_valid():
            user = User.objects.create_user(
                username=form.cleaned_data.get('username'),
                email=form.cleaned_data.get('email'),
                password=form.cleaned_data.get('password1')
            )

            Commercial.objects.create(
                user=user,
                name=form.cleaned_data.get('name'),
                email=form.cleaned_data.get('email'),
                phone=form.cleaned_data.get('phone'),
                status=form.cleaned_data.get('status'),
            )

            group = Group.objects.get(name="Commercial")
            user.groups.add(group)

            messages.success(request, f"Le commercial {user.username} a été créé avec succès !")
            return redirect('CommercialSuperAdmin')
        else:
            messages.error(request, "Formulaire invalide. Veuillez corriger les erreurs ci-dessous.")
    else:
        form = CommercialForm()
        
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'form': form,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/addCommercial.html', context)



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def updateCommercial(request, pk):
    commercial = get_object_or_404(Commercial, pk=pk)

    if request.method == 'POST':
        form = CommercialUpdateForm(request.POST, instance=commercial)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Commercial modifié avec succès.")
            return redirect('CommercialSuperAdmin')
        else:
            messages.error(request, "❌ Erreur dans le formulaire.")
    else:
        form = CommercialUpdateForm(instance=commercial)

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'form': form, 
        'commercial': commercial,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/updateCommercial.html', context)



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def deleteCommercial(request, pk):
    commercial = get_object_or_404(Commercial, pk=pk)

    if request.method == 'POST':
        if commercial.user:
            commercial.user.delete()
        commercial.delete()
        messages.success(request, "✅ Commercial supprimé avec succès.")
        return redirect('CommercialSuperAdmin')

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'commercial': commercial,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/deleteCommercial.html', context)



import csv
from django.http import HttpResponse
def export_commercials_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="commercials.csv"'

    writer = csv.writer(response)
    writer.writerow(['Nom', 'Email', 'Téléphone', 'Statut'])

    for c in Commercial.objects.all():
        writer.writerow([c.name, c.email, c.phone, c.status])

    return response




from django.http import HttpResponse
from openpyxl import Workbook

def export_commercials_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Commercials"

    ws.append(['Nom', 'Email', 'Téléphone', 'Statut'])

    for c in Commercial.objects.all():
        ws.append([c.name, c.email, c.phone, c.status])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=commercials.xlsx'
    wb.save(response)
    return response




def export_commercials_pdf(request):
    commercials = Commercial.objects.all()
    html = render_to_string("SuperAdmin/pdf_commercial_template.html", {
        'commercials': commercials,
        'now': timezone.now()
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="commercials.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Erreur lors de la génération du PDF")
    return response










#SuperAdmin can show details of the trace (image) of traceDemandeRecharger
@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def full_size_image_Super_Admin(request, traceDemandeRecharger_id):
    trace_demande_recharger = get_object_or_404(LaTraceDemandeRecharger, pk=traceDemandeRecharger_id)
    image = trace_demande_recharger.image
    cliente = trace_demande_recharger.cliente.user.username
    solde = trace_demande_recharger.solde
    updated_by = trace_demande_recharger.demande_recharger.updated_by
    status = trace_demande_recharger.demande_recharger.status
    motif = trace_demande_recharger.demande_recharger.motifNonAcceptation

    return render(request, 'SuperAdmin/full_size_image_Super_Admin.html', {'image': image, 'cliente': cliente, 'solde': solde,'updated_by':updated_by,'status':status,'motif':motif})


@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def all_demandes_recharger(request):
    # Filtres
    status = request.GET.get('status')
    type_demande = request.GET.get('type_demande')
    date = request.GET.get('date')
    code = request.GET.get('code')
    updated_by = request.GET.get('updated_by')
    per_page = int(request.GET.get('per_page', 10))
    page_number = request.GET.get('page')

    demandes = DemandeRecharger.objects.all().order_by('-date_created')

    if status:
        demandes = demandes.filter(status=status)
    if type_demande:
        demandes = demandes.filter(type_demande=type_demande)
    if date:
        demandes = demandes.filter(date_created__date=date)
    if code:
        demandes = demandes.filter(code_DemandeRecharger=code)
    if updated_by:
        demandes = demandes.filter(updated_by__id=updated_by)

    paginator = Paginator(demandes, per_page)
    page_obj = paginator.get_page(page_number)

    gestionnaires = GestionnaireComptes.objects.all()

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'demandes': page_obj.object_list,
        'page_obj': page_obj,
        'per_page': per_page,
        'status_filter': status,
        'type_demande': type_demande,
        'date': date,
        'code': code,
        'updated_by': updated_by,
        'gestionnaires': gestionnaires,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/all_demandes_recharger.html', context)


@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def detail_demande_recharge_superadmin(request, id):
    demande = get_object_or_404(
        DemandeRecharger.objects
            .select_related('cliente__user', 'updated_by__user'),
        id=id
    )
    traces_qs = demande.latracedemanderecharger_set.select_related(
        'updated_by__user'
    ).order_by('-date_created')

    # on transforme chaque trace en dict prêt à l’usage
    traces = [
        {
            "date": t.date_created.strftime("%d/%m/%Y %H:%M"),
            "gestionnaire": t.updated_by.user.username if t.updated_by else "—",
            "solde": f"{t.solde:.2f} MAD",
        }
        for t in traces_qs
    ]

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    contexte = {
        'demande': {
            'code': demande.code_DemandeRecharger,
            'client': demande.cliente.user.username,
            'montant': f"{demande.solde:.2f} MAD",
            'statut': demande.status,
            'date': demande.date_created.strftime("%d/%m/%Y %H:%M"),
            'gestionnaire': (
                demande.updated_by.user.username
                if demande.updated_by else '—'
            ),
            'motif': demande.motifNonAcceptation or '',
            'image_url': demande.image.url if demande.image else None,
        },
        'traces': traces,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    
    return render(request, 'SuperAdmin/detail_demande_recharge_superadmin.html', contexte)



def get_filtered_demandes(request):
    demandes = DemandeRecharger.objects.select_related('cliente', 'updated_by').all()
    status = request.GET.get('status')
    date = request.GET.get('date')
    code = request.GET.get('code')
    updated_by = request.GET.get('updated_by')

    if status:
        demandes = demandes.filter(status=status)
    if date:
        demandes = demandes.filter(date_created__date=date)
    if code:
        demandes = demandes.filter(code_DemandeRecharger__icontains=code)
    if updated_by:
        demandes = demandes.filter(updated_by__id=updated_by)

    return demandes.order_by('-date_created')




from xhtml2pdf import pisa
from django.template.loader import get_template
from django.http import HttpResponse
import io

def export_demandes_pdf(request):
    demandes = get_filtered_demandes(request)  # fonction à créer pour réutiliser les filtres
    template = get_template('SuperAdmin/demandes_recharger_pdf.html')
    html = template.render({'demandes': demandes})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="demandes.pdf"'
    pisa.CreatePDF(io.BytesIO(html.encode('UTF-8')), dest=response)
    return response



import csv

# def export_demandes_csv(request):
#     demandes = get_filtered_demandes(request)
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = 'attachment; filename="demandes.csv"'
#     writer = csv.writer(response)
#     writer.writerow(['Client', 'Solde', 'Statut', 'Motif', 'Date', 'Code', 'Mis à jour par'])
#     for d in demandes:
#         writer.writerow([
#             d.cliente.user.username,
#             d.solde,
#             d.status,
#             d.motifNonAcceptation,
#             d.date_created.strftime('%d/%m/%Y %H:%M'),
#             d.code_DemandeRecharger,
#             d.updated_by.user.username if d.updated_by else '—'
#         ])
#     return response



from openpyxl import Workbook

def export_demandes_excel(request):
    demandes = get_filtered_demandes(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Demandes de recharge"
    ws.append(['Client', 'Solde', 'Statut', 'Motif', 'Date', 'Code', 'Mis à jour par'])
    for d in demandes:
        ws.append([
            d.cliente.user.username,
            float(d.solde),
            d.status,
            d.motifNonAcceptation,
            d.date_created.strftime('%d/%m/%Y %H:%M'),
            d.code_DemandeRecharger,
            d.updated_by.user.username if d.updated_by else '—'
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="demandes.xlsx"'
    wb.save(response)
    return response





@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def gestionnaire_detail(request, gestionnaire_id):
    gestionnaire = get_object_or_404(GestionnaireComptes, id=gestionnaire_id)
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'gestionnaire': gestionnaire,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/gestionnaire_detail.html', context)






#List of LaTraceDemandeRecharger Demandes Recharge  [Done]
@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def traceDemandeRechargerDone(request): 
    traceDemandeRechargers = LaTraceDemandeRecharger.objects.filter(demande_recharger__status='Done').order_by('-date_created')
    context = {'traceDemandeRechargers': traceDemandeRechargers} 
    return render(request, "SuperAdmin/traceDemandeRechargerDone.html",context)




#List of LaTraceDemandeRecharger Demandes Recharge  [inacceptable]
@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def traceDemandeRechargerInacceptable(request): 
    traceDemandeRechargers = LaTraceDemandeRecharger.objects.filter(demande_recharger__status='inacceptable').order_by('-date_created')
    context = {'traceDemandeRechargers': traceDemandeRechargers} 
    return render(request, "SuperAdmin/traceDemandeRechargerInacceptable.html",context)





#SuperAdmin can show details of the trace (image) of traceDemandeRecharger
@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin'])
def full_size_image_NotDone_Super_Admin(request, DemandeRecharger_id):
    Demande_Recharger = get_object_or_404(DemandeRecharger, pk=DemandeRecharger_id)
    image = Demande_Recharger.image
    cliente = Demande_Recharger.cliente.user.username
    solde = Demande_Recharger.solde
    updated_by = Demande_Recharger.updated_by
    status = Demande_Recharger.status

    return render(request, 'SuperAdmin/full_size_image_Super_Admin.html', {'image': image, 'cliente': cliente, 'solde': solde,'updated_by':updated_by,'status':status})





@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def DemandeRechargerNotDone(request): 
    DemandeRechargerNotDones = DemandeRecharger.objects.filter(status='Not Done yet').order_by('-date_created')
    context = {'DemandeRechargerNotDones': DemandeRechargerNotDones} 
    return render(request, "SuperAdmin/DemandeRechargerNotDone.html",context)




@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def DemandeSupportAll(request): 
    DemandeSupports = DemandeSupport.objects.select_related('achat_support', 'updated_by', 'cliente')

    # Filtres
    status = request.GET.get('status')
    status_consome = request.GET.get('status_consome')
    updated_by = request.GET.get('updated_by')
    code = request.GET.get('code')
    cliente_username = request.GET.get('cliente_username')
    per_page = int(request.GET.get('per_page', 10))
    page_number = request.GET.get('page')

    # Application des filtres
    if status:
        DemandeSupports = DemandeSupports.filter(status=status)
    if status_consome:
        DemandeSupports = DemandeSupports.filter(achat_support__StatusConsomé=status_consome)
    if updated_by:
        DemandeSupports = DemandeSupports.filter(updated_by__id=updated_by)
    if code:
        DemandeSupports = DemandeSupports.filter(code_DemandeSupport__icontains=code)
    if cliente_username:
        DemandeSupports = DemandeSupports.filter(cliente__username=cliente_username)

    # Pagination
    paginator = Paginator(DemandeSupports.order_by('-date_created'), per_page)
    page_obj = paginator.get_page(page_number)

    techniciens = SupportTechnique.objects.filter(Status='Active')
    clients = User.objects.filter(groups__name='Client')  

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'DemandeSupports': page_obj.object_list,
        'page_obj': page_obj,
        'per_page': per_page,
        'status': status,
        'status_consome': status_consome,
        'updated_by': updated_by,
        'code': code,
        'cliente_username': cliente_username,
        'techniciens': techniciens,
        'clients': clients,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, "SuperAdmin/DemandeSupportAll.html", context)



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def detail_demande_support_superadmin(request, pk):
    demande = get_object_or_404(
        DemandeSupport.objects
            .select_related('cliente__user',
                            'achat_support__support',
                            'updated_by__user'),
        pk=pk
    )
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    return render(request, "SuperAdmin/detail_demande_support_superadmin.html", {
        'demande': demande,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    })




@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def supportTechnique_detail(request, supportTechnique_id):
    supportTechnique = get_object_or_404(SupportTechnique, id=supportTechnique_id)
    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'supportTechnique': supportTechnique,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/supportTechnique_detail.html', context)



from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import xlwt

def export_demande_support_filtered(request, format):
    demandes = DemandeSupport.objects.select_related('achat_support', 'updated_by', 'cliente')

    status = request.GET.get('status')
    if status and status != 'None':
        demandes = demandes.filter(status=status)

    status_consome = request.GET.get('status_consome')
    if status_consome and status_consome != 'None':
        demandes = demandes.filter(achat_support__StatusConsomé=status_consome)

    updated_by = request.GET.get('updated_by')
    if updated_by and updated_by != 'None' and updated_by.isdigit():
        demandes = demandes.filter(updated_by__id=int(updated_by))

    code = request.GET.get('code')
    if code and code != 'None':
        demandes = demandes.filter(code_DemandeSupport__icontains=code)

    cliente_username = request.GET.get('cliente_username')
    if cliente_username and cliente_username != 'None':
        demandes = demandes.filter(cliente__user__username=cliente_username)

    if format == 'excel':
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="demandes_support.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Demandes')
        row_num = 0
        columns = ['Client', 'Support', 'Prix', 'Status', 'Consommé', 'Technicien', 'Date', 'Code']
        for col_num, col in enumerate(columns):
            ws.write(row_num, col_num, col)
        for d in demandes:
            ws.write(row_num + 1, 0, d.cliente.user.username if d.cliente and d.cliente.user else "Sans nom")
            ws.write(row_num + 1, 1, d.achat_support.support.name if d.achat_support and d.achat_support.support else "")
            ws.write(row_num + 1, 2, float(d.achat_support.prix) if d.achat_support else 0)
            ws.write(row_num + 1, 3, d.status)
            ws.write(row_num + 1, 4, d.achat_support.StatusConsomé if d.achat_support else "")
            ws.write(row_num + 1, 5, d.updated_by.name if d.updated_by else "Non assigné")
            ws.write(row_num + 1, 6, d.date_created.strftime("%d/%m/%Y %H:%M"))
            ws.write(row_num + 1, 7, d.code_DemandeSupport)
            row_num += 1
        wb.save(response)
        return response

    elif format == 'pdf':
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        elements.append(Paragraph("📋 Liste des demandes de support", styles['Title']))
        elements.append(Spacer(1, 12))

        data = [
            ['Client', 'Support', 'Prix', 'Status', 'Consommé', 'Technicien', 'Date', 'Code']
        ]

        for d in demandes:
            data.append([
                d.cliente.user.username if d.cliente and d.cliente.user else "Sans nom",
                d.achat_support.support.name if d.achat_support and d.achat_support.support else "",
                f"{d.achat_support.prix} MAD" if d.achat_support else "",
                d.status,
                d.achat_support.StatusConsomé if d.achat_support else "",
                d.updated_by.name if d.updated_by else "Non assigné",
                d.date_created.strftime("%d/%m/%Y %H:%M"),
                d.code_DemandeSupport
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#d3d3d3')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.black),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ]))

        elements.append(table)
        doc.build(elements)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="demandes_support_table.pdf"'
        response.write(buffer.getvalue())
        buffer.close()
        return response

    return HttpResponse("Format non supporté", status=400)




#List of Demands Support Done [SuperAdmin]
@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def DemandeSupportDoneSA(request): 
    DemandeSupportsDone = DemandeSupport.objects.filter(status='Done').order_by('-date_created')
    # DemandeSupports = DemandeSupport.objects.order_by('-date_created')
    context = {'DemandeSupportsDone': DemandeSupportsDone} 
    return render(request, "SuperAdmin/DemandeSupportDoneSA.html",context)



#List of Demands Support Not Done yet [SuperAdmin]
@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def DemandeSupportNotDoneyetSA(request): 
    DemandeSupportsNotDoneyet = DemandeSupport.objects.filter(status='Not Done yet').order_by('-date_created')
    # DemandeSupports = DemandeSupport.objects.order_by('-date_created')
    context = {'DemandeSupportsNotDoneyet': DemandeSupportsNotDoneyet} 
    return render(request, "SuperAdmin/DemandeSupportNotDoneyetSA.html",context)




from django.utils.dateparse import parse_date

@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def history(request):
    model_name = request.GET.get('model_name')
    cliente_id = request.GET.get('cliente')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    per_page = int(request.GET.get('per_page', 10))
    page_number = request.GET.get('page')

    history_entries = History.objects.all().order_by('-date_created')

    if model_name:
        history_entries = history_entries.filter(model_name=model_name)
    if cliente_id:
        history_entries = history_entries.filter(cliente_id=cliente_id)
    if date_start:
        history_entries = history_entries.filter(date_created__date__gte=parse_date(date_start))
    if date_end:
        history_entries = history_entries.filter(date_created__date__lte=parse_date(date_end))

    paginator = Paginator(history_entries, per_page)
    page_obj = paginator.get_page(page_number)

    clientes = Cliente.objects.all()
    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    return render(request, 'SuperAdmin/history.html', {
        'history_entries': page_obj.object_list,
        'page_obj': page_obj,
        'per_page': per_page,
        'model_choices': History.MODEL_CHOICES,
        'clientes': clientes,
        'selected_model': model_name,
        'selected_cliente': cliente_id,
        'date_start': date_start,
        'date_end': date_end,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    })



from django.utils.dateparse import parse_date

def get_filtered_history(request):
    model_name = request.GET.get('model_name')
    cliente_id = request.GET.get('cliente')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')

    history_entries = History.objects.all().order_by('-date_created')

    if model_name:
        history_entries = history_entries.filter(model_name=model_name)
    if cliente_id:
        history_entries = history_entries.filter(cliente_id=cliente_id)
    if date_start:
        history_entries = history_entries.filter(date_created__date__gte=parse_date(date_start))
    if date_end:
        history_entries = history_entries.filter(date_created__date__lte=parse_date(date_end))

    return history_entries




from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
import io

def export_history_pdf(request):
    history_entries = get_filtered_history(request)
    template = get_template('SuperAdmin/history_pdf.html')
    html = template.render({'history_entries': history_entries})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="historique.pdf"'
    pisa.CreatePDF(io.BytesIO(html.encode('UTF-8')), dest=response)
    return response


from openpyxl import Workbook

def export_history_excel(request):
    history_entries = get_filtered_history(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Historique"
    ws.append(['ID', 'Type d’action', 'Instance ID', 'Client', 'Date'])

    for h in history_entries:
        ws.append([
            h.id,
            h.get_model_name_display(),
            h.instance_id,
            h.cliente.user.username if h.cliente else '—',
            h.date_created.strftime('%d/%m/%Y %H:%M')
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="historique.xlsx"'
    wb.save(response)
    return response



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def WebsitesListSuperAdmin(request):
    websites = Websites.objects.all()

    status = request.GET.get('status')
    catégorie = request.GET.get('catégorie')
    CMS = request.GET.get('CMS')
    langues = request.GET.get('langues')
    plan = request.GET.get('plan')

    if status and status != 'None':
        websites = websites.filter(status=status)
    if catégorie and catégorie != 'None':
        websites = websites.filter(catégorie=catégorie)
    if CMS and CMS != 'None':
        websites = websites.filter(CMS=CMS)
    if langues and langues != 'None':
        websites = websites.filter(langues=langues)
    if plan and plan != 'None':
        websites = websites.filter(plan=plan)

    websites = websites.order_by('-date_created')

    # ✅ Pagination
    paginator = Paginator(websites, 10)  # 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    
    context = {
        'page_obj': page_obj,
        'status': status,
        'catégorie': catégorie,
        'CMS': CMS,
        'langues': langues,
        'plan': plan,
        'catégories_list': ['Ecommerce','Blogs','Business','portfolio','Educational','News'],
        'cms_list': ['WordPress','Drupal'],
        'langues_list': ['Français','Anglais'],
        'plans_list': ['Free','Payant'],
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, "SuperAdmin/websites_list.html", context)




def get_filtered_websites(request):
    websites = Websites.objects.all()

    status = request.GET.get('status')
    catégorie = request.GET.get('catégorie')
    CMS = request.GET.get('CMS')
    langues = request.GET.get('langues')
    plan = request.GET.get('plan')

    if status and status != 'None':
        websites = websites.filter(status=status)
    if catégorie and catégorie != 'None':
        websites = websites.filter(catégorie=catégorie)
    if CMS and CMS != 'None':
        websites = websites.filter(CMS=CMS)
    if langues and langues != 'None':
        websites = websites.filter(langues=langues)
    if plan and plan != 'None':
        websites = websites.filter(plan=plan)

    return websites.order_by('-date_created')



from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
import io

def export_websites_pdf(request):
    websites = get_filtered_websites(request)
    template = get_template('SuperAdmin/websites_pdf.html')
    html = template.render({'websites': websites})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="websites.pdf"'
    pisa.CreatePDF(io.BytesIO(html.encode('UTF-8')), dest=response)
    return response




# import csv

# def export_websites_csv(request):
#     websites = get_filtered_websites(request)
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = 'attachment; filename="websites.csv"'
#     writer = csv.writer(response)
#     writer.writerow(['Nom', 'Catégorie', 'CMS', 'Langue', 'Forfait', 'Prix', 'Statut'])
#     for w in websites:
#         writer.writerow([
#             w.name,
#             w.catégorie,
#             w.CMS,
#             w.langues,
#             w.plan,
#             w.prix,
#             w.status
#         ])
#     return response



from openpyxl import Workbook

def export_websites_excel(request):
    websites = get_filtered_websites(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sites Web"
    ws.append(['Nom', 'Catégorie', 'CMS', 'Langue', 'Forfait', 'Prix', 'Statut'])
    for w in websites:
        ws.append([
            w.name,
            w.catégorie,
            w.CMS,
            w.langues,
            w.plan,
            float(w.prix),
            w.status
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="websites.xlsx"'
    wb.save(response)
    return response



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def add_website(request):
    if request.method == 'POST':
        form = WebsiteForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('websites_list_superadmin')
    else:
        form = WebsiteForm()
        
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    return render(request, 'SuperAdmin/add_website.html', {'form': form,'notifications': notifications,
        'messages_dropdown': messages_dropdown,})



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def website_details(request, id):
    website = get_object_or_404(Websites, id=id)
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    return render(request, 'SuperAdmin/website_details.html', {'website': website,'notifications': notifications,
        'messages_dropdown': messages_dropdown,})



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def edit_website(request, id):
    website = get_object_or_404(Websites, id=id)
    if request.method == 'POST':
        form = WebsiteForm(request.POST, request.FILES, instance=website)
        if form.is_valid():
            form.save()
            return redirect('websites_list_superadmin')
    else:
        form = WebsiteForm(instance=website)
        
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    return render(request, 'SuperAdmin/edit_website.html', {'form': form, 'website': website,'notifications': notifications,
        'messages_dropdown': messages_dropdown,})



def hide_website(request, id):
    website = get_object_or_404(Websites, id=id)
    website.is_visible = False
    website.save()
    return redirect('websites_list_superadmin')


from django.core.paginator import Paginator


@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def supports_list_superadmin(request):
    status = request.GET.get('status')

    supports = Supports.objects.all()
    if status:
        supports = supports.filter(status=status)

    supports = supports.order_by('-date_created')

    # ✅ Pagination
    paginator = Paginator(supports, 10)  # 10 éléments par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'page_obj': page_obj,  
        'status': status,
        'status_choices': ['Disponible', 'No Disponible'],
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/supports_list.html', context)



def get_filtered_supports(request):
    supports = Supports.objects.all()
    status = request.GET.get('status')
    if status and status != 'None':
        supports = supports.filter(status=status)
    return supports.order_by('-date_created')





def export_supports_pdf(request):
    supports = get_filtered_supports(request)
    template = get_template('SuperAdmin/supports_pdf.html')
    html = template.render({'supports': supports})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="supports.pdf"'
    pisa.CreatePDF(io.BytesIO(html.encode('UTF-8')), dest=response)
    return response



def export_supports_csv(request):
    supports = get_filtered_supports(request)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="supports.csv"'
    writer = csv.writer(response)
    writer.writerow(['Nom', 'Description', 'Prix', 'Date de création', 'Statut'])
    for s in supports:
        writer.writerow([
            s.name,
            s.description,
            s.prix,
            s.date_created.strftime('%d/%m/%Y'),
            s.status
        ])
    return response



def export_supports_excel(request):
    supports = get_filtered_supports(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Supports"
    ws.append(['Nom', 'Description', 'Prix', 'Date de création', 'Statut'])
    for s in supports:
        ws.append([
            s.name,
            s.description,
            float(s.prix),
            s.date_created.strftime('%d/%m/%Y'),
            s.status
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="supports.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def add_support(request):
    if request.method == 'POST':
        form = SupportForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('supports_list_superadmin')  
    else:
        form = SupportForm()
        
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    return render(request, 'SuperAdmin/add_support.html', {'form': form,'notifications': notifications,
        'messages_dropdown': messages_dropdown,})



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def support_details(request, id):
    support = get_object_or_404(Supports, id=id)
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    return render(request, 'SuperAdmin/support_details.html', {'support': support,'notifications': notifications,
        'messages_dropdown': messages_dropdown,})



@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def edit_support(request, id):
    support = get_object_or_404(Supports, id=id)
    if request.method == 'POST':
        form = SupportForm(request.POST, instance=support)
        if form.is_valid():
            form.save()
            return redirect('supports_list_superadmin')  
    else:
        form = SupportForm(instance=support)
        
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    return render(request, 'SuperAdmin/edit_support.html', {'form': form, 'support': support,'notifications': notifications,
        'messages_dropdown': messages_dropdown,})




@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def hide_support(request, id):
    support = get_object_or_404(Supports, id=id)
    support.status = 'No Disponible'
    support.save()
    return redirect('supports_list_superadmin')





@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def tickets_list(request):
    tickets = Ticket.objects.all()

    # 🔍 Recherche rapide
    search_query = request.GET.get('search')
    if search_query:
        tickets = tickets.filter(
            Q(code_Ticket__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(typeTicket__icontains=search_query) |
            Q(Branche__icontains=search_query) |
            Q(cliente__user__username__icontains=search_query)
        )

    # ✅ Filtres
    status = request.GET.get('status')
    typeTicket = request.GET.get('typeTicket')
    branche = request.GET.get('Branche')
    cliente_id = request.GET.get('cliente')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')

    if status and status != 'None':
        tickets = tickets.filter(status=status)
    if typeTicket and typeTicket != 'None':
        tickets = tickets.filter(typeTicket=typeTicket)
    if branche and branche != 'None':
        tickets = tickets.filter(Branche=branche)
    if cliente_id and cliente_id != 'None':
        tickets = tickets.filter(cliente_id=cliente_id)
    if date_start and date_end:
        tickets = tickets.filter(date_created__date__range=[date_start, date_end])

    # 🔃 Tri
    sort_by = request.GET.get('sort_by')
    if sort_by == 'date_asc':
        tickets = tickets.order_by('date_created')
    elif sort_by == 'date_desc':
        tickets = tickets.order_by('-date_created')
    elif sort_by == 'status_asc':
        tickets = tickets.order_by('status')
    elif sort_by == 'status_desc':
        tickets = tickets.order_by('-status')
    else:
        tickets = tickets.order_by('-date_created')

    # ✅ Pagination dynamique
    per_page = request.GET.get('per_page')
    try:
        per_page = int(per_page)
        if per_page not in [10, 25, 50, 100]:
            per_page = 10
    except (TypeError, ValueError):
        per_page = 10

    paginator = Paginator(tickets, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()
    
    context = {
        'page_obj': page_obj,
        'per_page': per_page,
        'status_choices': Ticket.STATUS_CHOICES,
        'type_choices': Ticket.objects.values_list('typeTicket', flat=True).distinct(),
        'branche_choices': Ticket.objects.values_list('Branche', flat=True).distinct(),
        'clientes': Ticket.objects.values_list('cliente__id', 'cliente__user__username').distinct(),
        'status': status,
        'typeTicket': typeTicket,
        'branche': branche,
        'cliente_id': cliente_id,
        'date_start': date_start,
        'date_end': date_end,
        'search_query': search_query,
        'sort_by': sort_by,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/tickets_list.html', context)






@login_required(login_url='login')
@allowedUsers(allowedGroups=['SuperAdmin']) 
def ticket_detail(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    conversations = ticket.conversations.order_by('timestamp')

    user_type = None
    user_id = None

    try:
        cliente = Cliente.objects.get(user=request.user)
        user_type = 'Cliente'
        user_id = cliente.id
    except Cliente.DoesNotExist:
        try:
            support = SupportTechnique.objects.get(user=request.user)
            user_type = 'SupportTechnique'
            user_id = support.id
        except SupportTechnique.DoesNotExist:
            try:
                gestionnaire = GestionnaireComptes.objects.get(user=request.user)
                user_type = 'GestionnaireComptes'
                user_id = gestionnaire.id
            except GestionnaireComptes.DoesNotExist:
                pass
             
    notifications = get_superadmin_notifications()
    messages_dropdown = get_all_ticket_messages()            
    
    context = {
        'ticket': ticket,
        'conversations': conversations,
        'user_type': user_type,
        'user_id': user_id,
        'notifications': notifications,
        'messages_dropdown': messages_dropdown,
    }
    return render(request, 'SuperAdmin/ticket_detail.html', context)



from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import io

def ticket_pdf(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    conversations = ticket.conversations.order_by('timestamp')

    template_path = 'SuperAdmin/detail_ticket_pdf.html'
    context = {'ticket': ticket, 'conversations': conversations}
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ticket_{ticket.code_Ticket}.pdf"'

    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response.write(result.getvalue())
        return response
    else:
        return HttpResponse("Erreur lors de la génération du PDF", status=500)
